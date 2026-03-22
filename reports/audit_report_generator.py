"""
Audit Report Generator — produces a professional Excel audit summary report.
Follows financial industry color-coding standards.
Blue = inputs/hardcoded, Black = formulas, color-coded severity bands.
"""

import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

DB_PATH = os.environ.get("DB_PATH", "db/equity_qa.db")

# ── Color palette ──────────────────────────────────────────────────────────────
C_HEADER_BG    = "1B2A4A"   # deep navy
C_HEADER_FG    = "FFFFFF"
C_TITLE_BG     = "0F3460"
C_ACCENT       = "E94560"   # red accent
C_CRITICAL_BG  = "FDECEA"
C_CRITICAL_FG  = "B71C1C"
C_HIGH_BG      = "FFF3E0"
C_HIGH_FG      = "E65100"
C_MEDIUM_BG    = "FFFDE7"
C_MEDIUM_FG    = "F57F17"
C_GOOD_BG      = "E8F5E9"
C_GOOD_FG      = "1B5E20"
C_ROW_ALT      = "F7F9FC"
C_BORDER       = "D0D7E3"
C_SECTION_BG   = "EEF2F7"
C_BLUE_INPUT   = "0000FF"   # industry standard: hardcoded inputs
C_BLACK_CALC   = "000000"   # industry standard: formulas/calcs


def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def bottom_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(bottom=s)


def header_font(size=11, bold=True, color=C_HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)


def body_font(size=10, bold=False, color=C_BLACK_CALC):
    return Font(name="Arial", size=size, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ── Sheet 1: Cover / Summary ───────────────────────────────────────────────────
def build_cover(ws, summary: dict):
    ws.title = "Audit Summary"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 3

    # ── Title block ──
    ws.merge_cells("B1:F1")
    ws["B1"] = "EQUITY DATA QUALITY AUDIT REPORT"
    ws["B1"].font = Font("Arial", size=16, bold=True, color=C_HEADER_FG)
    ws["B1"].fill = fill(C_TITLE_BG)
    ws["B1"].alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("B2:F2")
    ws["B2"] = f"Public Ownership Domain  |  Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws["B2"].font = Font("Arial", size=10, color="AABBCC")
    ws["B2"].fill = fill(C_HEADER_BG)
    ws["B2"].alignment = center()
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10

    # ── KPI cards row ──
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 24
    ws.row_dimensions[7].height = 10

    kpis = [
        ("Total Records",   summary["total_records"],    "0",    C_HEADER_BG,  C_HEADER_FG),
        ("Open Defects",    summary["open_defects"],     "0",    C_ACCENT,     C_HEADER_FG),
        ("Defect Rate",     summary["defect_rate_pct"],  "0.0%", "D84040",     C_HEADER_FG),
        ("Critical",        summary["critical_count"],   "0",    C_CRITICAL_FG, C_HEADER_FG),
        ("Rules Applied",   12,                          "0",    "1B5E20",     C_HEADER_FG),
    ]

    kpi_cols = ["B", "C", "D", "E", "F"]
    for col, (label, value, fmt, bg, fg) in zip(kpi_cols, kpis):
        ws.merge_cells(f"{col}4:{col}4")
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = Font("Arial", size=9, bold=True, color="AABBDD")
        ws[f"{col}4"].fill = fill(C_HEADER_BG)
        ws[f"{col}4"].alignment = center()

        ws[f"{col}5"] = value if fmt != "0.0%" else value / 100
        ws[f"{col}5"].font = Font("Arial", size=20, bold=True, color=C_HEADER_FG)
        ws[f"{col}5"].fill = fill(bg)
        ws[f"{col}5"].alignment = center()
        ws[f"{col}5"].number_format = fmt

        ws[f"{col}6"] = ""
        ws[f"{col}6"].fill = fill(bg)

    # ── Section: Defects by Severity ──
    ws.row_dimensions[8].height = 20
    ws.merge_cells("B8:F8")
    ws["B8"] = "  DEFECT BREAKDOWN BY SEVERITY"
    ws["B8"].font = header_font(10, True, C_HEADER_FG)
    ws["B8"].fill = fill(C_HEADER_BG)
    ws["B8"].alignment = left()

    headers = ["Severity", "Open Defects", "% of Total", "Typical Impact", "Recommended SLA"]
    cols = ["B", "C", "D", "E", "F"]
    ws.row_dimensions[9].height = 18
    for col, h in zip(cols, headers):
        ws[f"{col}9"] = h
        ws[f"{col}9"].font = header_font(9, True, "CCDDEE")
        ws[f"{col}9"].fill = fill(C_HEADER_BG)
        ws[f"{col}9"].alignment = center()
        ws[f"{col}9"].border = thin_border()

    severity_data = [
        ("Critical", summary["critical_count"], "Block client delivery",        "24 hours",   C_CRITICAL_BG, C_CRITICAL_FG),
        ("High",     summary["high_count"],      "Affects downstream analytics", "72 hours",   C_HIGH_BG,     C_HIGH_FG),
        ("Medium",   summary["medium_count"],    "Data quality risk",            "2 weeks",    C_MEDIUM_BG,   C_MEDIUM_FG),
    ]

    total_defects = summary["open_defects"] or 1
    for i, (sev, count, impact, sla, bg, fg) in enumerate(severity_data, start=10):
        ws.row_dimensions[i].height = 20
        row_data = [sev, count, f"=C{i}/C12", impact, sla]
        for col, val in zip(cols, row_data):
            cell = ws[f"{col}{i}"]
            cell.value = val
            cell.fill = fill(bg)
            cell.border = thin_border()
            cell.alignment = center() if col != "B" else left()
            if col == "B":
                cell.font = Font("Arial", size=10, bold=True, color=fg)
            elif col == "D":
                cell.number_format = "0.0%"
                cell.font = body_font(10, False, fg)
            else:
                cell.font = body_font(10, False, fg)

    # Totals row
    ws.row_dimensions[12].height = 20
    totals = ["TOTAL", f"=SUM(C10:C11)", "", "", ""]
    for col, val in zip(cols, totals):
        ws[f"{col}12"] = val
        ws[f"{col}12"].font = Font("Arial", size=10, bold=True, color=C_HEADER_FG)
        ws[f"{col}12"].fill = fill(C_HEADER_BG)
        ws[f"{col}12"].alignment = center()
        ws[f"{col}12"].border = thin_border()

    # Fix total formula - hardcode for simplicity
    ws["C12"] = summary["open_defects"]
    ws["C12"].font = Font("Arial", size=10, bold=True, color=C_HEADER_FG)
    ws["C12"].fill = fill(C_HEADER_BG)
    ws["C12"].alignment = center()

    # Fix percentage formulas
    for i, (_, count, _, _, _, _) in enumerate(severity_data, start=10):
        ws[f"D{i}"] = count / total_defects
        ws[f"D{i}"].number_format = "0.0%"
        ws[f"D{i}"].fill = fill(severity_data[i-10][4])
        ws[f"D{i}"].border = thin_border()
        ws[f"D{i}"].alignment = center()
        ws[f"D{i}"].font = body_font(10, False, severity_data[i-10][5])

    ws.row_dimensions[13].height = 10

    # ── Section: Audit Metadata ──
    ws.row_dimensions[14].height = 20
    ws.merge_cells("B14:F14")
    ws["B14"] = "  AUDIT METADATA"
    ws["B14"].font = header_font(10, True, C_HEADER_FG)
    ws["B14"].fill = fill(C_HEADER_BG)
    ws["B14"].alignment = left()

    meta = [
        ("Last Audit Run",       summary.get("latest_run_at", "N/A")[:19] if summary.get("latest_run_at") else "N/A"),
        ("Records Scanned",      f"{summary['total_records']:,}"),
        ("Rules Applied",        "12"),
        ("Pipeline Version",     "1.0.0"),
        ("Data Domain",          "Public Ownership — Shareholding"),
        ("Methodology",          "Rule-based + Cross-record validation"),
        ("Report Generated By",  "Equity QA Pipeline API"),
    ]

    for i, (label, value) in enumerate(meta, start=15):
        ws.row_dimensions[i].height = 18
        ws.merge_cells(f"B{i}:C{i}")
        ws[f"B{i}"] = label
        ws[f"B{i}"].font = body_font(10, True, "444466")
        ws[f"B{i}"].fill = fill(C_SECTION_BG)
        ws[f"B{i}"].alignment = left()
        ws[f"B{i}"].border = bottom_border()

        ws.merge_cells(f"D{i}:F{i}")
        ws[f"D{i}"] = value
        ws[f"D{i}"].font = Font("Arial", size=10, color=C_BLUE_INPUT)
        ws[f"D{i}"].fill = fill("FFFFFF")
        ws[f"D{i}"].alignment = left()
        ws[f"D{i}"].border = bottom_border()


# ── Sheet 2: Defect Log ────────────────────────────────────────────────────────
def build_defect_log(ws):
    ws.title = "Defect Log"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols_config = [
        ("Defect ID",        8),
        ("Record ID",        12),
        ("Company",          8),
        ("Rule ID",          8),
        ("Rule Name",        30),
        ("Severity",         10),
        ("Field Affected",   14),
        ("Field Value",      18),
        ("Status",           10),
        ("Detected At",      18),
        ("Resolution Notes", 30),
    ]

    for i, (header, width) in enumerate(cols_config, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font(9, True, C_HEADER_FG)
        cell.fill = fill(C_HEADER_BG)
        cell.alignment = center()
        cell.border = thin_border()

    ws.row_dimensions[1].height = 20

    conn = get_connection()
    rows = conn.execute(
        """SELECT defect_id, record_id, company_ticker, rule_id, rule_name,
                  severity, field_affected, field_value, status, detected_at, resolution_notes
           FROM defect_log ORDER BY
             CASE severity WHEN 'Critical' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
             company_ticker"""
    ).fetchall()
    conn.close()

    SEV_STYLE = {
        "Critical": (C_CRITICAL_BG, C_CRITICAL_FG),
        "High":     (C_HIGH_BG,     C_HIGH_FG),
        "Medium":   (C_MEDIUM_BG,   C_MEDIUM_FG),
    }

    for r_idx, row in enumerate(rows, start=2):
        row_bg = C_ROW_ALT if r_idx % 2 == 0 else "FFFFFF"
        sev = row["severity"]
        sev_bg, sev_fg = SEV_STYLE.get(sev, ("FFFFFF", "000000"))

        values = [
            row["defect_id"], row["record_id"], row["company_ticker"],
            row["rule_id"], row["rule_name"], row["severity"],
            row["field_affected"], row["field_value"] or "—",
            row["status"], row["detected_at"][:10] if row["detected_at"] else "",
            row["resolution_notes"] or "—",
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border()
            cell.alignment = center() if c_idx in [1, 3, 4, 6, 9] else left()

            if c_idx == 6:  # Severity column
                cell.fill = fill(sev_bg)
                cell.font = Font("Arial", size=9, bold=True, color=sev_fg)
            else:
                cell.fill = fill(row_bg)
                cell.font = body_font(9)

        ws.row_dimensions[r_idx].height = 16

    # Auto-filter
    ws.auto_filter.ref = f"A1:K{len(rows)+1}"


# ── Sheet 3: Rule Performance ──────────────────────────────────────────────────
def build_rule_performance(ws):
    ws.title = "Rule Performance"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for col, width in zip(["B","C","D","E","F","G"], [10, 35, 12, 14, 20, 3]):
        ws.column_dimensions[col].width = width

    ws.merge_cells("B1:F1")
    ws["B1"] = "VALIDATION RULE PERFORMANCE"
    ws["B1"].font = header_font(12, True, C_HEADER_FG)
    ws["B1"].fill = fill(C_TITLE_BG)
    ws["B1"].alignment = center()
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 8

    headers = ["Rule ID", "Rule Name", "Severity", "Defects Found", "Companies Affected", "Preventive Action"]
    for i, h in enumerate(["B","C","D","E","F"], start=0):
        if i < len(headers):
            ws[f"{h}3"] = headers[i]
            ws[f"{h}3"].font = header_font(9, True, C_HEADER_FG)
            ws[f"{h}3"].fill = fill(C_HEADER_BG)
            ws[f"{h}3"].alignment = center()
            ws[f"{h}3"].border = thin_border()
    ws.row_dimensions[3].height = 18

    conn = get_connection()
    rule_stats = conn.execute(
        """SELECT rule_id, rule_name, severity,
                  COUNT(*) as defect_count,
                  COUNT(DISTINCT company_ticker) as companies
           FROM defect_log WHERE status='Open'
           GROUP BY rule_id ORDER BY
             CASE severity WHEN 'Critical' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
             defect_count DESC"""
    ).fetchall()
    conn.close()

    from pipeline.defect_classifier import PREVENTIVE_ACTIONS
    SEV_STYLE = {
        "Critical": (C_CRITICAL_BG, C_CRITICAL_FG),
        "High":     (C_HIGH_BG,     C_HIGH_FG),
        "Medium":   (C_MEDIUM_BG,   C_MEDIUM_FG),
    }

    for r_idx, row in enumerate(rule_stats, start=4):
        sev_bg, sev_fg = SEV_STYLE.get(row["severity"], ("FFFFFF", "000000"))
        row_bg = C_ROW_ALT if r_idx % 2 == 0 else "FFFFFF"

        data = [
            ("B", row["rule_id"],   center(), body_font(9, True)),
            ("C", row["rule_name"], left(),   body_font(9)),
            ("D", row["severity"],  center(), Font("Arial", size=9, bold=True, color=sev_fg)),
            ("E", row["defect_count"], center(), body_font(9, True)),
        ]
        for col, val, align, font in data:
            cell = ws[f"{col}{r_idx}"]
            cell.value = val
            cell.font = font
            cell.alignment = align
            cell.border = thin_border()
            cell.fill = fill(sev_bg if col == "D" else row_bg)

        # Companies affected with mini bar representation
        ws[f"E{r_idx}"] = row["defect_count"]
        ws[f"E{r_idx}"].font = body_font(9, True)
        ws[f"E{r_idx}"].alignment = center()
        ws[f"E{r_idx}"].fill = fill(row_bg)
        ws[f"E{r_idx}"].border = thin_border()

        ws.row_dimensions[r_idx].height = 18

    # Add chart data for bar chart (severity counts)
    chart_start = len(rule_stats) + 6
    ws[f"B{chart_start}"] = "Severity"
    ws[f"C{chart_start}"] = "Count"
    ws[f"B{chart_start}"].font = body_font(9, True)
    ws[f"C{chart_start}"].font = body_font(9, True)

    conn = get_connection()
    sev_counts = conn.execute(
        "SELECT severity, COUNT(*) as cnt FROM defect_log WHERE status='Open' GROUP BY severity"
    ).fetchall()
    conn.close()

    sev_map = {r["severity"]: r["cnt"] for r in sev_counts}
    for i, sev in enumerate(["Critical", "High", "Medium"], start=1):
        ws.cell(row=chart_start+i, column=2, value=sev)
        ws.cell(row=chart_start+i, column=3, value=sev_map.get(sev, 0))

    chart = BarChart()
    chart.type = "col"
    chart.title = "Open Defects by Severity"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.style = 10
    chart.width = 14
    chart.height = 10

    data_ref = Reference(ws, min_col=3, min_row=chart_start, max_row=chart_start+3)
    cats = Reference(ws, min_col=2, min_row=chart_start+1, max_row=chart_start+3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"B{chart_start+5}")


# ── Sheet 4: Company Risk Profiles ────────────────────────────────────────────
def build_risk_profiles(ws):
    ws.title = "Company Risk Profiles"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    for col, w in zip(["B","C","D","E","F","G","H","I"], [14,14,12,10,10,10,14,3]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:H1")
    ws["B1"] = "COMPANY RISK PROFILES"
    ws["B1"].font = header_font(12, True, C_HEADER_FG)
    ws["B1"].fill = fill(C_TITLE_BG)
    ws["B1"].alignment = center()
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8

    headers = ["Company", "Risk Label", "Risk Score", "Critical", "High", "Medium", "Total Defects"]
    for i, (col, h) in enumerate(zip(["B","C","D","E","F","G","H"], headers)):
        ws[f"{col}3"] = h
        ws[f"{col}3"].font = header_font(9, True, C_HEADER_FG)
        ws[f"{col}3"].fill = fill(C_HEADER_BG)
        ws[f"{col}3"].alignment = center()
        ws[f"{col}3"].border = thin_border()
    ws.row_dimensions[3].height = 18

    conn = get_connection()
    rows = conn.execute(
        """SELECT company_ticker,
                  COUNT(*) AS total,
                  SUM(CASE WHEN severity='Critical' THEN 1 ELSE 0 END) AS crit,
                  SUM(CASE WHEN severity='High' THEN 1 ELSE 0 END) AS high,
                  SUM(CASE WHEN severity='Medium' THEN 1 ELSE 0 END) AS med
           FROM defect_log WHERE status='Open'
           GROUP BY company_ticker
           ORDER BY crit DESC, high DESC, total DESC"""
    ).fetchall()
    conn.close()

    RISK_STYLE = {
        "High Risk":   ("FDECEA", "B71C1C"),
        "Medium Risk": ("FFF3E0", "E65100"),
        "Low Risk":    ("E8F5E9", "1B5E20"),
    }

    for r_idx, row in enumerate(rows, start=4):
        score = row["crit"]*10 + row["high"]*6 + row["med"]*3
        label = "High Risk" if score >= 20 else "Medium Risk" if score >= 8 else "Low Risk"
        risk_bg, risk_fg = RISK_STYLE[label]
        row_bg = C_ROW_ALT if r_idx % 2 == 0 else "FFFFFF"

        vals = [row["company_ticker"], label, score, row["crit"], row["high"], row["med"], row["total"]]
        for col, val in zip(["B","C","D","E","F","G","H"], vals):
            cell = ws[f"{col}{r_idx}"]
            cell.value = val
            cell.alignment = center()
            cell.border = thin_border()
            if col == "C":
                cell.fill = fill(risk_bg)
                cell.font = Font("Arial", size=9, bold=True, color=risk_fg)
            elif col == "E":
                cell.fill = fill(row_bg)
                cell.font = Font("Arial", size=9, bold=True, color=C_CRITICAL_FG if row["crit"] > 0 else C_BLACK_CALC)
            else:
                cell.fill = fill(row_bg)
                cell.font = body_font(9)
        ws.row_dimensions[r_idx].height = 17


# ── Main entry point ───────────────────────────────────────────────────────────
def generate_audit_report(output_path: str = "reports/audit_summary.xlsx") -> str:
    os.makedirs("reports", exist_ok=True)

    conn = get_connection()
    total = conn.execute("SELECT COUNT(*) FROM shareholding_raw").fetchone()[0]
    latest_run = conn.execute("SELECT run_at FROM audit_runs ORDER BY run_id DESC LIMIT 1").fetchone()
    sev = conn.execute(
        "SELECT severity, COUNT(*) as cnt FROM defect_log WHERE status='Open' GROUP BY severity"
    ).fetchall()
    open_total = conn.execute("SELECT COUNT(*) FROM defect_log WHERE status='Open'").fetchone()[0]
    resolved = conn.execute("SELECT COUNT(*) FROM defect_log WHERE status='Resolved'").fetchone()[0]
    conn.close()

    sev_map = {r["severity"]: r["cnt"] for r in sev}
    summary = {
        "total_records":    total,
        "latest_run_at":    latest_run["run_at"] if latest_run else None,
        "open_defects":     open_total,
        "defect_rate_pct":  round(open_total / total * 100, 2) if total else 0,
        "critical_count":   sev_map.get("Critical", 0),
        "high_count":       sev_map.get("High", 0),
        "medium_count":     sev_map.get("Medium", 0),
        "resolved_count":   resolved,
    }

    wb = Workbook()

    # Build all sheets
    build_cover(wb.active, summary)

    ws_log = wb.create_sheet()
    build_defect_log(ws_log)

    ws_rules = wb.create_sheet()
    build_rule_performance(ws_rules)

    ws_risk = wb.create_sheet()
    build_risk_profiles(ws_risk)

    wb.save(output_path)
    print(f"Report saved → {output_path}")
    return output_path


if __name__ == "__main__":
    generate_audit_report()
